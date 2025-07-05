from sqlalchemy.orm import Session
from . import models, schemas
from minio import Minio
import pytesseract
from PIL import Image
from io import BytesIO
import spacy
import pdfplumber
import openpyxl
from datetime import datetime
import re
from typing import Dict, Any, List, Optional

# Načtení rozšířeného modelu spaCy
try:
    nlp = spacy.load("cs_core_news_lg")
except OSError:
    # Pokud model není nainstalován, stáhneme ho
    spacy.cli.download("cs_core_news_lg")
    nlp = spacy.load("cs_core_news_lg")

def extract_dates(text: str) -> List[datetime]:
    """Extrahuje data z textu pomocí regulárních výrazů a spaCy"""
    dates = []
    # Regex pro různé formáty dat
    date_patterns = [
        r'\d{1,2}\.\s*\d{1,2}\.\s*\d{4}',
        r'\d{4}-\d{1,2}-\d{1,2}',
    ]
    
    for pattern in date_patterns:
        matches = re.finditer(pattern, text)
        for match in matches:
            try:
                date_str = match.group()
                if '.' in date_str:
                    date = datetime.strptime(date_str, '%d.%m.%Y')
                else:
                    date = datetime.strptime(date_str, '%Y-%m-%d')
                dates.append(date)
            except ValueError:
                continue
    
    return dates

def extract_measurements(text: str) -> List[Dict[str, Any]]:
    """Extrahuje rozměry a další měřitelné hodnoty z textu"""
    measurements = []
    doc = nlp(text)
    
    # Hledání číselných hodnot s jednotkami
    measurement_patterns = [
        (r'(\d+(?:,\d+)?)\s*(mm|cm|m|kg|t)', 'dimension'),
        (r'(\d+(?:,\d+)?)\s*(kg/m²|t/m²)', 'load_capacity'),
    ]
    
    for pattern, measure_type in measurement_patterns:
        matches = re.finditer(pattern, text)
        for match in matches:
            value, unit = match.groups()
            value = float(value.replace(',', '.'))
            measurements.append({
                'type': measure_type,
                'value': value,
                'unit': unit
            })
    
    return measurements

def extract_milestones(text: str) -> List[Dict[str, Any]]:
    """Extrahuje milníky projektu z textu"""
    doc = nlp(text)
    milestones = []
    
    # Klíčová slova pro identifikaci milníků
    milestone_keywords = ['milník', 'fáze', 'etapa', 'deadline', 'termín']
    
    for sent in doc.sents:
        sent_text = sent.text.lower()
        if any(keyword in sent_text for keyword in milestone_keywords):
            dates = extract_dates(sent.text)
            if dates:
                milestones.append({
                    'description': sent.text,
                    'date': dates[0]
                })
    
    return milestones

def extract_key_data_from_text(text: str) -> Dict[str, Any]:
    """
    Hlavní funkce pro extrakci klíčových dat z textu.
    Zpracovává text pomocí spaCy a dalších metod pro získání strukturovaných dat.
    """
    doc = nlp(text)
    
    # Inicializace výsledného slovníku
    extracted_data = {
        'dates': extract_dates(text),
        'measurements': extract_measurements(text),
        'milestones': extract_milestones(text),
        'entities': {},
        'keywords': set()
    }
    
    # Extrakce pojmenovaných entit
    for ent in doc.ents:
        if ent.label_ not in extracted_data['entities']:
            extracted_data['entities'][ent.label_] = []
        extracted_data['entities'][ent.label_].append({
            'text': ent.text,
            'label': ent.label_
        })
    
    # Extrakce klíčových slov (podstatná jména a jejich fráze)
    for chunk in doc.noun_chunks:
        if len(chunk.text.split()) > 1:  # Pouze víceslovné fráze
            extracted_data['keywords'].add(chunk.text)
    
    # Převod množiny na seznam pro JSON serializaci
    extracted_data['keywords'] = list(extracted_data['keywords'])
    
    return extracted_data

def process_pdf_document(file_content: bytes) -> str:
    """Zpracování PDF dokumentu a extrakce textu"""
    text = ""
    with pdfplumber.open(BytesIO(file_content)) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text

def process_xlsx_document(file_content: bytes) -> str:
    """Zpracování Excel dokumentu a extrakce textu"""
    text = []
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            row_text = " ".join(str(cell.value) for cell in row if cell.value)
            if row_text:
                text.append(row_text)
    return "\n".join(text)

def get_project(db: Session, project_id: int):
    return db.query(models.Project).filter(models.Project.id == project_id).first()

def get_projects(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Project).offset(skip).limit(limit).all()

def create_project(db: Session, project: schemas.ProjectCreate):
    db_project = models.Project(**project.dict(), owner_id=1) # Hardcoded owner_id
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project

def update_project(db: Session, project_id: int, project: schemas.ProjectCreate):
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if db_project:
        db_project.name = project.name
        db_project.description = project.description
        db.commit()
        db.refresh(db_project)
    return db_project

def delete_project(db: Session, project_id: int):
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if db_project:
        db.delete(db_project)
        db.commit()
    return db_project

def get_documents(db: Session, project_id: int, category: str | None = None, skip: int = 0, limit: int = 100):
    query = db.query(models.Document).filter(models.Document.project_id == project_id)
    if category:
        query = query.filter(models.Document.category == category)
    return query.offset(skip).limit(limit).all()

def create_progress_log(db: Session, progress_log: schemas.ProgressLogCreate, project_id: int):
    db_progress_log = models.ProgressLog(**progress_log.dict(), project_id=project_id)
    db.add(db_progress_log)
    db.commit()
    db.refresh(db_progress_log)
    return db_progress_log

def update_progress_log(db: Session, progress_log_id: int, progress_log: schemas.ProgressLogCreate):
    db_progress_log = db.query(models.ProgressLog).filter(models.ProgressLog.id == progress_log_id).first()
    if db_progress_log:
        db_progress_log.date = progress_log.date
        db_progress_log.percentage_completed = progress_log.percentage_completed
        db_progress_log.notes = progress_log.notes
        db.commit()
        db.refresh(db_progress_log)
    return db_progress_log

def delete_progress_log(db: Session, progress_log_id: int):
    db_progress_log = db.query(models.ProgressLog).filter(models.ProgressLog.id == progress_log_id).first()
    if db_progress_log:
        db.delete(db_progress_log)
        db.commit()
    return db_progress_log

def get_progress_logs(db: Session, project_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.ProgressLog).filter(models.ProgressLog.project_id == project_id).offset(skip).limit(limit).all()

def get_project_overall_progress(db: Session, project_id: int):
    progress_logs = db.query(models.ProgressLog).filter(models.ProgressLog.project_id == project_id).all()
    if not progress_logs:
        return 0
    total_percentage = sum([log.percentage_completed for log in progress_logs])
    return total_percentage / len(progress_logs)

def perform_ocr_on_document(db: Session, document_id: int, minio_client: Minio):
    """Rozšířená funkce pro OCR a extrakci dat z různých typů dokumentů"""
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        return None, None

    bucket_name = "ranger-bucket"
    try:
        response = minio_client.get_object(bucket_name, db_document.filename)
        file_content = response.read()
        
        # Zpracování podle typu dokumentu
        if db_document.filename.lower().endswith('.pdf'):
            text = process_pdf_document(file_content)
        elif db_document.filename.lower().endswith(('.xlsx', '.xls')):
            text = process_xlsx_document(file_content)
        else:
            # Pro obrázky použijeme původní OCR
            image = Image.open(BytesIO(file_content))
            text = pytesseract.image_to_string(image, lang='ces')
        
        # Extrakce klíčových dat
        extracted_data = extract_key_data_from_text(text)
        return text, extracted_data
        
    except Exception as e:
        print(f"Error processing document: {e}")
        return None, None

def get_total_projects_count(db: Session):
    return db.query(models.Project).count()

def get_completed_projects_count(db: Session):
    completed_projects = db.query(models.Project).join(models.ProgressLog).filter(models.ProgressLog.percentage_completed == 100).distinct().count()
    return completed_projects


    total_overall_progress = 0
    for project in all_projects:
        total_overall_progress += get_project_overall_progress(db, project.id)
    return total_overall_progress / len(all_projects)

# --- Detekce anomálií ve fotodokumentaci ---
import cv2
import numpy as np

def detect_anomaly_in_image(document_id: int, minio_client: Minio):
    """
    Skutečná logika detekce anomálií ve fotodokumentaci pomocí OpenCV.
    Detekuje základní typy anomálií: chybějící/nové objekty, poškození, barevné odchylky.
    """
    bucket_name = "ranger-bucket"
    from .models import SessionLocal, Document
    db = SessionLocal()
    db_document = db.query(Document).filter(Document.id == document_id).first()
    if not db_document:
        return {"anomaly_detected": False, "message": "Dokument nenalezen."}
    try:
        response = minio_client.get_object(bucket_name, db_document.filename)
        file_content = response.read()
        image = Image.open(BytesIO(file_content)).convert('RGB')
        img_np = np.array(image)
        img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)

        # 1. Detekce rozmazání (ostrost)
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        blurry = laplacian_var < 50  # Prahová hodnota pro rozmazání

        # 2. Detekce dominantní barvy (např. příliš červené = rez)
        mean_color = cv2.mean(img_cv)
        red_ratio = mean_color[2] / (mean_color[0] + mean_color[1] + mean_color[2] + 1e-5)
        rusty = red_ratio > 0.5

        # 3. Detekce poškození (velké tmavé oblasti)
        _, thresh = cv2.threshold(gray, 40, 255, cv2.THRESH_BINARY_INV)
        dark_area = np.sum(thresh == 255) / (thresh.shape[0] * thresh.shape[1])
        damaged = dark_area > 0.15

        # 4. (Volitelně) Detekce chybějících objektů pomocí jednoduché kontury
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        large_contours = [cnt for cnt in contours if cv2.contourArea(cnt) > 5000]
        missing_parts = len(large_contours) < 2  # Očekáváme aspoň 2 velké objekty

        # Výsledek
        anomaly_types = []
        if blurry:
            anomaly_types.append("Rozmazaný snímek")
        if rusty:
            anomaly_types.append("Podezření na rez (převaha červené)")
        if damaged:
            anomaly_types.append("Velké tmavé oblasti - možné poškození")
        if missing_parts:
            anomaly_types.append("Chybějící části/objekty")

        if anomaly_types:
            return {
                "anomaly_detected": True,
                "message": "Byly detekovány anomálie.",
                "details": ", ".join(anomaly_types)
            }
        else:
            return {
                "anomaly_detected": False,
                "message": "Žádné zjevné anomálie nebyly detekovány. (Omezená přesnost)",
                "details": None
            }
    except Exception as e:
        return {"anomaly_detected": False, "message": f"Chyba při zpracování: {e}"}
